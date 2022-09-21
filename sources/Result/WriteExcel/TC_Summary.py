from openpyxl import load_workbook
import xlsxwriter as xw
from decimal import Decimal

#--------------TC01-----------
count_pass01 = 0
count_fail01 = 0
TC01 = load_workbook(filename='TC01_Search_Survey_result.xlsx')
Sheet_TC01 = TC01['TestData']
for i in range(2, 9):
    Result = Sheet_TC01['F'+str(i)].value
    if Result == 'PASS':
        count_pass01 += 1
    else:
        count_fail01 += 1
print('TC01')
print('PASS : ', count_pass01)
print('FAIL : ', count_fail01)
Sum_Result01 = count_pass01+count_fail01
print('Sum_Result : ', Sum_Result01)
print('------------------------------')
#--------------TC02-----------
count_pass02 = 0
count_fail02 = 0
TC02 = load_workbook(filename='TC02_Login_result.xlsx')
Sheet_TC02 = TC02['TestData']
for i in range(2, 24):
    Result = Sheet_TC02['F'+str(i)].value
    if Result == 'PASS':
        count_pass02 += 1
    else:
        count_fail02 += 1
print('TC02')
print('PASS : ', count_pass02)
print('FAIL : ', count_fail02)
Sum_Result02 = count_pass02+count_fail02
print('Sum_Result : ', Sum_Result02)
print('------------------------------')
#--------------TC03-----------
count_pass03 = 0
count_fail03 = 0
TC03 = load_workbook(filename='TC03_Register_result.xlsx')
Sheet_TC03 = TC03['TestData']
for i in range(2, 73):
    Result = Sheet_TC03['L'+str(i)].value
    if Result == 'PASS':
        count_pass03 += 1
    else:
        count_fail03 += 1
print('TC03')
print('PASS : ', count_pass03)
print('FAIL : ', count_fail03)
Sum_Result03 = count_pass03+count_fail03
print('Sum_Result : ', Sum_Result03)
print('------------------------------')
#--------------TC04-----------
workbook = xw.Workbook('./Test_Summary.xlsx')
worksheet = workbook.add_worksheet('TC_Summary')
f2 = workbook.add_format({})
worksheet.set_column('A:F', 20)
bold = workbook.add_format({'bold': True, 'font_size': '15', 'center_across': True, 'bg_color': '#FFCC00', 'border': 2, 'border_color': 'black'})
text = workbook.add_format({'font_size': '15', 'center_across': True, 'border': 2, 'border_color': 'black'})
worksheet.write('A1', 'TestCase', bold)
worksheet.write('B1', 'TestData', bold)
worksheet.write('C1', 'Pass', bold)
worksheet.write('D1', 'Fail', bold)
worksheet.write('E1', '%Pass', bold)
worksheet.write('F1', '%Fail', bold)

n = 22
for i in range(2, n):
    worksheet.write('A'+str(i), 'TC0'+str(i-1), text)
    worksheet.write('B2', count_pass01+count_fail01, text)
    worksheet.write('C2', count_pass01, text)
    worksheet.write('D2', count_fail01, text)
    worksheet.write('E2', round(Decimal((Sum_Result01-count_fail01)*100/Sum_Result01), 2), text)
    worksheet.write('F2', round(Decimal((Sum_Result01-count_pass01)*100/Sum_Result01), 2), text)

worksheet.write('B3', count_pass02+count_fail02, text)
worksheet.write('C3', count_pass02, text)
worksheet.write('D3', count_fail02, text)
worksheet.write('E3', round(Decimal((Sum_Result02-count_fail02)*100/Sum_Result02), 2), text)
worksheet.write('F3', round(Decimal((Sum_Result02-count_pass02)*100/Sum_Result02), 2), text)
workbook.close()

# TC03 = pd.read_excel("TC03_Register_result.xlsx")
# print("read successfully TC03")
#
# TC05 = pd.read_excel("TC05_Add_Questions_result.xlsx")
# print("read successfully TC05")
#
# TC07 = pd.read_excel("TC07_Send_Email_Complaint_result.xlsx")
# print("read successfully TC07")
#
# TC08 = pd.read_excel("TC08_Add_Survey_result.xlsx")
# print("read successfully TC08")
#
# TC10 = pd.read_excel("TC10_Add_Answer_result.xlsx")
# print("read successfully TC10")
#
# TC11 = pd.read_excel("TC11_Remove_Questions_result.xlsx")
# print("read successfully TC11")
#
# TC12 = pd.read_excel("TC12_Add_News_result.xlsx")
# print("read successfully TC12")
#
# TC14 = pd.read_excel("TC02_Login_result.xlsx")
# print("read successfully TC14")
#
# TC15 = pd.read_excel("TC15_Upload_File_result.xlsx")
# print("read successfully TC15")
#
# TC16 = pd.read_excel("TC16_Remove_File_result.xlsx")
# print("read successfully TC16")
#
# TC17 = pd.read_excel("TC17_Add_Surveyor_result.xlsx")
# print("read successfully TC17")
#
# TC19 = pd.read_excel("TC19_Add_Officer_result.xlsx")
# print("read successfully TC19")


